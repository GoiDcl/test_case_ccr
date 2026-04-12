import os
from tempfile import NamedTemporaryFile

from django import forms
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.forms import Form, FileField
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect
from django.urls import re_path
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from config.settings import BASE_DIR
from event_app.constants import EventStatusChoices
from event_app.models import Event, EventImage, EventLocation, WeatherForecast
from event_app.tasks import import_excel_events_file_task


class EventImageInline(admin.StackedInline):
    model = EventImage
    extra = 0


class WeatherForecastInline(admin.StackedInline):
    model = WeatherForecast
    extra = 0


class EventLocationForm(forms.ModelForm):
    def clean(self):
        cleaned_data = super().clean()
        try:
            longitude, latitude = cleaned_data['coordinates'].split(';')
            longitude = float(longitude)
            latitude = float(latitude)
        except ValueError:
            self.add_error('coordinates', 'Неправильный формат координат')
        return cleaned_data

    class Meta:
        model = EventLocation
        fields = "__all__"


@admin.register(EventLocation)
class EventLocationAdmin(admin.ModelAdmin):
    # model = EventLocation
    form = EventLocationForm
    list_display = ('name', 'coordinates')
    search_fields = ('name', 'coordinates')
    inlines = (WeatherForecastInline,)


@admin.register(Event)
class EventAdmin(admin.ModelAdmin):
    model = Event
    list_display = (
        'name',
        'published_at',
        'starting_at',
        'finishing_at',
        'creator',
        'location__name',
        'status',
    )
    search_fields = ('name', 'location__name')
    list_filter = ('published_at', 'starting_at', 'finishing_at', 'location__name', 'rating')
    inlines = (EventImageInline,)
    actions = ('export_as_xlsx',)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related('location')

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            re_path(r'^import_events/$', self.import_events),
        ]
        return my_urls + urls

    def export_as_xlsx(self, request, queryset):
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=events.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'Events'

        columns = [
            ('ID', 10),
            ('Название', 30),
            ('Место проведения', 25),
            ('Описание', 40),
            ('Дата публикации', 20),
            ('Начало', 20),
            ('Завершение', 20),
            ('Автор', 20),
            ('Рейтинг', 10),
            ('Статус', 15),
        ]

        row_num = 1
        for col_num, (title, width) in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = title
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = width

        for obj in queryset:
            row_num += 1
            row = [
                obj.id,
                obj.name,
                obj.location.name if obj.location else '',
                (obj.description or '').strip(),
                obj.published_at.strftime('%Y-%m-%d %H:%M') if obj.published_at else '',
                obj.starting_at.strftime('%Y-%m-%d %H:%M') if obj.starting_at else '',
                obj.finishing_at.strftime('%Y-%m-%d %H:%M') if obj.finishing_at else '',
                obj.creator or '',
                obj.rating or '',
                dict(EventStatusChoices.choices).get(obj.status),
            ]
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = str(value)
                cell.alignment = Alignment(wrap_text=True)

        wb.save(response)
        return response

    export_as_xlsx.short_description = 'Экспортировать в .xlsx'

    class ExcelFileForm(Form):
        excel_file = FileField()

    @transaction.atomic
    def import_events(self, request):
        if not request.user.is_superuser:
            raise Http404
        if request.method == 'POST':
            excel_file = request.FILES['excel_file']
            # TODO: celery игнорит файл или ловит ошибку?
            # with NamedTemporaryFile(dir=BASE_DIR, delete=False, suffix='.xlsx') as temp_file:
            #     for chunk in excel_file.chunks():
            #         temp_file.write(chunk)
            #     temp_path = temp_file.name
            # import_excel_events_file_task.delay(temp_path)
            import_excel_events_file_task(excel_file)
            self.message_user(request, 'Файл с мероприятиями отправлен в обработку.')
            return redirect('.')

        context = {'form': self.ExcelFileForm()}
        return render(request, 'admin/upload_events.html', context=context)
