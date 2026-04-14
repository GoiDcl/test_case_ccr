import logging
from random import randint, choice

from celery import shared_task
from django.conf import settings
from django.utils import timezone
from openpyxl.reader.excel import load_workbook

from event_app.constants import WindDirectionChoices, EventStatusChoices
from event_app.models import Event, EventLocation, WeatherForecast


logger = logging.getLogger('celery')


@shared_task(bind=True, max_retries=settings.MAX_RETRY_COUNT, default_retry_delay=settings.MAX_RETRY_DELAY)
def import_excel_events_file_task(self, excel_file):
    """Загрузка файла Excel с мероприятиями."""
    try:
        workbook = load_workbook(excel_file, read_only=True)
        wb = workbook.active
        start_cell = wb.cell(row=1, column=1)
        events_dict = {}
        locations_dict = {}

        for row in wb.iter_rows(min_row=start_cell.row + 1):
            event_name = row[start_cell.column - 1].value
            description = row[start_cell.column].value
            published = row[start_cell.column + 1].value
            starting = row[start_cell.column + 2].value
            finishing = row[start_cell.column + 3].value
            location_name = row[start_cell.column + 4].value
            coordinates = row[start_cell.column + 5].value
            rating = row[start_cell.column + 6].value

            if event_name and event_name != 'Название':
                locations_dict[location_name] = {
                    'name': location_name,
                    'coordinates': coordinates,
                }
                events_dict[event_name] = {
                    'name': event_name,
                    'description': description,
                    'published_at': published,
                    'starting_at': starting,
                    'finishing_at': finishing,
                    'rating': rating,
                    'location_name': location_name,
                }
        existing_locations = {
            location.name: location
            for location in EventLocation.objects.all()
        }
        locations_to_create = [
            EventLocation(**data)
            for name, data in locations_dict.items()
            if name not in existing_locations
        ]
        if locations_to_create:
            created_locations = EventLocation.objects.bulk_create(locations_to_create)
            existing_locations.update({location.name: location for location in created_locations})
            logger.info(f'Создано новых мест проведения мероприятий: {len(created_locations)}')

        existing_events = set(
            Event.objects.filter(name__in=events_dict).values_list('name', flat=True)
        )
        events_to_create = []
        for name, data in events_dict.items():
            if name in existing_events:
                continue
            location_name = data.pop('location_name')
            location = existing_locations.get(location_name)
            events_to_create.append(Event(**data, location=location))
        if events_to_create:
            new_events = Event.objects.bulk_create(events_to_create)
            logger.info(f'Создано новых мероприятий: {len(new_events)}')

    except Exception as exception:
        logger.exception(f'Возникла ошибка при импорте мероприятий: {exception}')
        raise self.retry(exc=exception)
    finally:
        import os
        os.remove(excel_file)


@shared_task(bind=True, max_retries=settings.MAX_RETRY_COUNT, default_retry_delay=settings.MAX_RETRY_DELAY)
def check_event_statuses(self):
    """Проверка и обновление статусов мероприятий, рассылка при начале мероприятия."""
    events = Event.objects.filter(status=EventStatusChoices.PENDING)
    events_to_update = []
    events_to_notify = []
    for event in events:
        if event.starting_at < timezone.now():
            event.status = EventStatusChoices.GOING
            events_to_update.append(event)
            events_to_notify.append(event)
        if event.finishing_at < timezone.now():
            event.status = EventStatusChoices.ENDED
            events_to_update.append(event)
    if events_to_update:
        updated = Event.objects.bulk_update(events_to_update, ['status'])
        logger.info(f'Обновлено мероприятий: {updated}')
    if events_to_notify:
        try:
            for event in events_to_notify:
                event.send_email()
        except Exception as exception:
            logger.exception(f'Ошибка при рассылке: {exception}')
            raise self.retry(exc=exception)


@shared_task(bind=True, max_retries=settings.MAX_RETRY_COUNT, default_retry_delay=settings.MAX_RETRY_DELAY)
def weather_forecasts_update_task(self):
    """Периодическое обновление прогноза погоды в местах проведения мероприятий."""
    locations = EventLocation.objects.all()
    fields = ['temperature', 'humidity', 'atmospheric_pressure', 'wind_direction', 'wind_speed']
    forecasts = []
    new_forecasts = []
    for location in locations:
        temperature = randint(-90, 60)
        humidity = randint(0, 100)
        atmospheric_pressure = randint(650, 820)
        wind_direction = choice(WindDirectionChoices.values)
        wind_speed = randint(0, 100)
        try:
            forecast = location.weather
            location.weather.temperature = temperature
            location.weather.humidity = humidity
            location.weather.atmospheric_pressure = atmospheric_pressure
            location.weather.wind_direction = wind_direction
            location.weather.wind_speed = wind_speed
            forecasts.append(forecast)
        except EventLocation.weather.RelatedObjectDoesNotExist:
            new_forecast = WeatherForecast(
                location=location,
                temperature=temperature,
                humidity=humidity,
                atmospheric_pressure=atmospheric_pressure,
                wind_direction=wind_direction,
                wind_speed=wind_speed,
            )
            new_forecasts.append(new_forecast)

    if forecasts:
        try:
            updated = WeatherForecast.objects.bulk_update(forecasts, fields)
            logger.info(f'Обновлено прогнозов погоды: {updated}')
        except Exception as exception:
            logger.exception(f'Ошибка при обновлении прогнозов погоды: {exception}')
            raise self.retry(exc=exception)
    if new_forecasts:
        try:
            created = WeatherForecast.objects.bulk_create(new_forecasts)
            logger.info(f'Создано прогнозов погоды: {len(created)}')
        except Exception as exception:
            logger.exception(f'Ошибка при создании прогнозов погоды: {exception}')
            raise self.retry(exc=exception)
